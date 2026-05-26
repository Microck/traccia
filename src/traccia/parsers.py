from __future__ import annotations

import csv
import gzip
import io
import json
import mimetypes
import re
import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from urllib.parse import ParseResult, urlparse

from traccia.config import TracciaConfig
from traccia.document_normalizer import DOCUMENT_SOURCE_TYPES, normalize_document
from traccia.family_normalizer import (
    collect_attachment_references_from_html,
    normalize_family_content,
)
from traccia.models import (
    AttachmentKind,
    ParsedDocument,
    ParsedSpan,
    Sensitivity,
    SourceAttachment,
    SourceCategory,
    SourceDocument,
    SourceFamily,
    SourceStatus,
    SourceType,
)
from traccia.utils import file_sha256, now_utc, short_hash, source_id_for_relative_path

SUPPORTED_EXTENSIONS = {
    ".md": SourceType.MARKDOWN,
    ".markdown": SourceType.MARKDOWN,
    ".txt": SourceType.TEXT,
    ".log": SourceType.TEXT,
    ".html": SourceType.TEXT,
    ".htm": SourceType.TEXT,
    ".xml": SourceType.TEXT,
    ".svg": SourceType.TEXT,
    ".yaml": SourceType.TEXT,
    ".yml": SourceType.TEXT,
    ".toml": SourceType.TEXT,
    ".ini": SourceType.TEXT,
    ".properties": SourceType.TEXT,
    ".cfg": SourceType.TEXT,
    ".conf": SourceType.TEXT,
    ".env": SourceType.TEXT,
    ".jsonl": SourceType.TEXT,
    ".json5": SourceType.TEXT,
    ".ndjson": SourceType.TEXT,
    ".py": SourceType.CODE,
    ".js": SourceType.CODE,
    ".ts": SourceType.CODE,
    ".tsx": SourceType.CODE,
    ".rs": SourceType.CODE,
    ".go": SourceType.CODE,
    ".sql": SourceType.CODE,
    ".sh": SourceType.CODE,
    ".bash": SourceType.CODE,
    ".zsh": SourceType.CODE,
    ".json": SourceType.JSON,
    ".csv": SourceType.CSV,
    ".xlsx": SourceType.SPREADSHEET,
    ".ics": SourceType.CALENDAR,
    ".mbox": SourceType.TEXT,
    ".pdf": SourceType.PDF,
    ".docx": SourceType.DOCX,
}

COMPRESSED_TEXT_INNER_EXTENSIONS = {
    ".jsonl",
    ".ndjson",
    ".txt",
    ".log",
    ".md",
    ".markdown",
}

IMAGE_SOURCE_EXTENSIONS = {
    ".avif",
    ".bmp",
    ".gif",
    ".heic",
    ".heif",
    ".jpeg",
    ".jpg",
    ".png",
    ".webp",
}

KNOWN_BINARY_EXTENSIONS = {
    ".7z",
    ".aac",
    ".avif",
    ".avi",
    ".bmp",
    ".deb",
    ".dmg",
    ".exe",
    ".flac",
    ".gif",
    ".gz",
    ".heic",
    ".heif",
    ".ico",
    ".jar",
    ".jpeg",
    ".jpg",
    ".m4a",
    ".mkv",
    ".mov",
    ".mp3",
    ".mp4",
    ".m2ts",
    ".ogg",
    ".otf",
    ".osz",
    ".png",
    ".psd",
    ".rar",
    ".rpm",
    ".ttf",
    ".tgz",
    ".wav",
    ".webm",
    ".webp",
    ".wmv",
    ".woff",
    ".woff2",
    ".mid",
}

TEXT_READ_ENCODINGS = (
    "utf-8",
    "utf-8-sig",
    "utf-16",
    "utf-16-le",
    "utf-16-be",
    "latin-1",
)

MAX_UNKNOWN_EXTENSION_TEXT_SNIFF_BYTES = 768 * 1024

USER_AGENT_LOG_ROLES = {"user", "human", "developer"}
NON_USER_AGENT_LOG_ROLES = {
    "assistant",
    "ai",
    "agent",
    "model",
    "system",
    "tool",
    "function",
    "thinking",
    "reasoning",
    "analysis",
}



@dataclass(slots=True)
class ParsedSourceContent:
    text: str
    spans: list[ParsedSpan]
    source_type: SourceType
    source_category: SourceCategory
    parser: str
    title: str | None
    created_at: datetime | None
    metadata: dict[str, Any]
    attachments: list[SourceAttachment]


def supported_file(path: Path) -> bool:
    return (
        path.suffix.lower() in SUPPORTED_EXTENSIONS
        or _compressed_text_inner_suffix(path) in SUPPORTED_EXTENSIONS
        or _looks_like_extensionless_text_export(path)
    )


def ingestable_file(path: Path) -> bool:
    return supported_file(path) or sniff_text_file(path)


def detect_source_type(path: Path) -> SourceType:
    source_type = SUPPORTED_EXTENSIONS.get(_compressed_text_inner_suffix(path) or path.suffix.lower())
    if source_type:
        return source_type
    if path.suffix.lower() in IMAGE_SOURCE_EXTENSIONS:
        return SourceType.IMAGE
    if _looks_like_extensionless_text_export(path):
        return SourceType.TEXT
    if sniff_text_file(path):
        return SourceType.TEXT
    supported_suffixes = ", ".join(sorted(SUPPORTED_EXTENSIONS))
    raise ValueError(
        f"Unsupported source type for {path}. "
        f"Supported suffixes: {supported_suffixes}; extensionless text exports are sniffed when safe."
    )


def _looks_like_extensionless_text_export(path: Path) -> bool:
    if path.suffix:
        return False
    return "/google ai studio/" in path.as_posix().lower()


def sniff_text_file(path: Path) -> bool:
    if _compressed_text_inner_suffix(path):
        try:
            with gzip.open(path, "rb") as handle:
                return sniff_text_bytes(handle.read(4096))
        except OSError:
            return False
    if path.suffix.lower() in KNOWN_BINARY_EXTENSIONS:
        return False
    if not path.suffix and _file_too_large_for_unknown_text_sniff(path):
        return False
    try:
        with path.open("rb") as handle:
            return sniff_text_bytes(handle.read(4096))
    except OSError:
        return False


def _file_too_large_for_unknown_text_sniff(path: Path) -> bool:
    try:
        return path.stat().st_size > MAX_UNKNOWN_EXTENSION_TEXT_SNIFF_BYTES
    except OSError:
        return True


def sniff_text_bytes(sample: bytes) -> bool:
    if not sample:
        return True
    if sample.startswith((b"\xef\xbb\xbf", b"\xff\xfe", b"\xfe\xff")):
        return True
    if b"\x00" in sample:
        return False
    try:
        sample.decode("utf-8")
        return True
    except UnicodeDecodeError:
        pass
    try:
        sample.decode("utf-16")
        return True
    except UnicodeDecodeError:
        pass

    decoded = sample.decode("latin-1")
    printable = sum(character.isprintable() or character.isspace() for character in decoded)
    return printable / len(decoded) >= 0.85


def parse_document(
    path: Path,
    *,
    project_relative_path: Path,
    config: TracciaConfig | None = None,
    source_family: SourceFamily | None = None,
    source_family_subproduct: str | None = None,
) -> ParsedDocument:
    source_id = source_id_for_relative_path(project_relative_path)
    parsed_source = _parse_source_content(
        path=path,
        project_relative_path=project_relative_path,
        source_id=source_id,
        config=config,
        source_family=source_family,
        source_family_subproduct=source_family_subproduct,
    )
    source = SourceDocument(
        source_id=source_id,
        uri=path.resolve().as_uri(),
        source_type=parsed_source.source_type,
        source_category=parsed_source.source_category,
        parser=parsed_source.parser,
        sha256=file_sha256(path),
        created_at=parsed_source.created_at or datetime_or_none(path.stat().st_mtime),
        ingested_at=now_utc(),
        title=parsed_source.title or path.stem.replace("-", " ").replace("_", " ").title(),
        language=_guess_language(path),
        sensitivity=Sensitivity.PRIVATE,
        metadata={
            "relative_import_path": project_relative_path.as_posix(),
            "filename": path.name,
            **parsed_source.metadata,
            "classification_reason": _classification_reason(
                path=path,
                project_relative_path=project_relative_path,
                source_category=parsed_source.source_category,
            ),
        },
        status=SourceStatus.ACTIVE,
    )
    return ParsedDocument(
        source=source,
        text=parsed_source.text,
        spans=parsed_source.spans,
        attachments=parsed_source.attachments,
    )


def _parse_source_content(
    *,
    path: Path,
    project_relative_path: Path,
    source_id: str,
    config: TracciaConfig | None = None,
    source_family: SourceFamily | None = None,
    source_family_subproduct: str | None = None,
) -> ParsedSourceContent:
    source_type = detect_source_type(path)
    attachment_references: list[dict[str, Any]] = []
    family_normalized = normalize_family_content(
        path=path,
        project_relative_path=project_relative_path,
        source_type=source_type,
        source_family=source_family,
        source_family_subproduct=source_family_subproduct,
    )
    if family_normalized:
        text = family_normalized.text
        parser = family_normalized.parser
        metadata = family_normalized.metadata
        title = family_normalized.title
        created_at = family_normalized.created_at
        attachment_references = list(metadata.get("attachment_references") or [])
    if source_type == SourceType.JSON:
        if not family_normalized:
            payload = json.loads(_read_text_with_fallback(path))
            attachment_references = _collect_json_attachment_references(payload)
            structured = _parse_structured_json_source(
                payload=payload,
                path=path,
                project_relative_path=project_relative_path,
                source_id=source_id,
            )
            if structured:
                remote_media_references = _collect_remote_media_url_references_from_text(structured.text)
                combined_attachment_references = attachment_references + remote_media_references
                structured.attachments = _build_linked_attachments(
                    source_path=path,
                    attachment_references=combined_attachment_references,
                    config=config,
                )
                if combined_attachment_references:
                    structured.metadata["attachment_reference_count"] = len(combined_attachment_references)
                if remote_media_references:
                    structured.metadata["remote_media_reference_count"] = len(remote_media_references)
                if structured.attachments:
                    structured.metadata["attachment_count"] = len(structured.attachments)
                return structured
            text = json.dumps(payload, indent=2, sort_keys=True)
            parser = _parser_name(source_type)
            metadata = {}
            title = None
            created_at = None
    elif source_type in DOCUMENT_SOURCE_TYPES:
        if not family_normalized:
            normalized = normalize_document(path, source_type=source_type, config=config)
            text = normalized.text
            parser = normalized.parser
            metadata = normalized.metadata
            title = None
            created_at = None
    elif not family_normalized:
        text = _read_text(path, source_type)
        if source_type == SourceType.TEXT and path.suffix.lower() in {".html", ".htm"}:
            attachment_references = collect_attachment_references_from_html(text)
        parser = _parser_name(source_type)
        metadata = {}
        title = None
        created_at = None

    source_category = _classify_source_category(
        path=path,
        project_relative_path=project_relative_path,
        source_type=source_type,
        text=text,
    )
    agent_log = _parse_agent_log_text(
        text=text,
        source_id=source_id,
        project_relative_path=project_relative_path,
        source_category=source_category,
    )
    metadata = dict(metadata)
    if agent_log:
        text, agent_log_spans, agent_log_metadata = agent_log
        metadata.update(agent_log_metadata)
        parser = agent_log_metadata["structured_export_kind"]
        source_type = SourceType.CHAT
        source_category = SourceCategory.AI_DIALOGUE
    remote_media_references = _collect_remote_media_url_references_from_text(text)
    combined_attachment_references = attachment_references + remote_media_references
    attachments = _build_linked_attachments(
        source_path=path,
        attachment_references=combined_attachment_references,
        config=config,
    )
    if combined_attachment_references:
        metadata["attachment_reference_count"] = len(combined_attachment_references)
    if remote_media_references:
        metadata["remote_media_reference_count"] = len(remote_media_references)
    if attachments:
        metadata["attachment_count"] = len(attachments)
    return ParsedSourceContent(
        text=text,
        spans=agent_log_spans if agent_log else _segment_text(text=text, source_id=source_id),
        source_type=source_type,
        source_category=source_category,
        parser=parser,
        title=title,
        created_at=created_at,
        metadata=metadata,
        attachments=attachments,
    )


def _read_text(path: Path, source_type: SourceType) -> str:
    if source_type == SourceType.IMAGE:
        return f"Image file: {path.name}"
    if source_type in {SourceType.MARKDOWN, SourceType.TEXT, SourceType.CODE, SourceType.CALENDAR}:
        return _read_text_with_fallback(path)
    if source_type == SourceType.JSON:
        payload = json.loads(_read_text_with_fallback(path))
        return json.dumps(payload, indent=2, sort_keys=True)
    if source_type == SourceType.CSV:
        csv_text = _read_text_with_fallback(path)
        rows: list[str] = []
        reader = csv.DictReader(io.StringIO(csv_text))
        for row in reader:
            rows.append(", ".join(f"{key}={value}" for key, value in row.items()))
        return "\n".join(rows)
    if source_type == SourceType.SPREADSHEET:
        return _read_spreadsheet_text(path)
    supported_types = ", ".join(item.value for item in SourceType)
    raise ValueError(
        f"Unsupported source type {source_type!r} for {path}. Supported source types: {supported_types}."
    )


def _read_spreadsheet_text(path: Path) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        blocks: list[str] = []
        for worksheet in workbook.worksheets:
            rows: list[str] = []
            for row in worksheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                if values:
                    rows.append(" | ".join(values))
            if rows:
                blocks.append(f"# Sheet: {worksheet.title}\n" + "\n".join(rows))
        return "\n\n".join(blocks)
    finally:
        workbook.close()


def _read_text_with_fallback(path: Path) -> str:
    if _compressed_text_inner_suffix(path):
        raw = gzip.decompress(path.read_bytes())
        for encoding in TEXT_READ_ENCODINGS:
            try:
                return raw.decode(encoding)
            except UnicodeDecodeError:
                continue
        return raw.decode("utf-8", errors="replace")
    for encoding in TEXT_READ_ENCODINGS:
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def _parser_name(source_type: SourceType) -> str:
    return {
        SourceType.MARKDOWN: "markdown",
        SourceType.TEXT: "text",
        SourceType.CODE: "code",
        SourceType.JSON: "json",
        SourceType.CSV: "csv",
        SourceType.SPREADSHEET: "xlsx",
        SourceType.PDF: "pypdf",
        SourceType.DOCX: "python-docx",
        SourceType.CHAT: "chat",
        SourceType.BOOKMARKS: "bookmarks",
        SourceType.CALENDAR: "calendar",
        SourceType.ISSUE_TRACKER: "issue-tracker",
        SourceType.PORTFOLIO: "portfolio",
        SourceType.SLIDES: "slides",
        SourceType.IMAGE: "image",
    }[source_type]


def _guess_language(path: Path) -> str | None:
    suffix = _compressed_text_inner_suffix(path) or path.suffix.lower()
    return {
        ".py": "python",
        ".js": "javascript",
        ".ts": "typescript",
        ".tsx": "tsx",
        ".rs": "rust",
        ".go": "go",
        ".sql": "sql",
        ".sh": "shell",
        ".bash": "shell",
        ".zsh": "shell",
        ".md": "markdown",
        ".markdown": "markdown",
        ".json": "json",
        ".yaml": "yaml",
        ".yml": "yaml",
        ".toml": "toml",
        ".html": "html",
        ".htm": "html",
        ".xml": "xml",
        ".csv": "csv",
    }.get(suffix)


def _compressed_text_inner_suffix(path: Path) -> str | None:
    suffixes = [suffix.lower() for suffix in path.suffixes]
    if len(suffixes) < 2 or suffixes[-1] != ".gz":
        return None
    inner_suffix = suffixes[-2]
    if inner_suffix in COMPRESSED_TEXT_INNER_EXTENSIONS:
        return inner_suffix
    return None


def _segment_text(*, text: str, source_id: str) -> list[ParsedSpan]:
    spans: list[ParsedSpan] = []
    if not text.strip():
        return spans

    cursor = 0
    for block in [segment.strip() for segment in text.split("\n\n") if segment.strip()]:
        block_start = text.find(block, cursor)
        block_heading = block.splitlines()[0].lstrip("# ").strip() if block.startswith("#") else None
        inner_cursor = block_start
        for line in [item.strip() for item in block.splitlines() if item.strip()]:
            start = text.find(line, inner_cursor)
            end = start + len(line)
            line_start = text[:start].count("\n") + 1
            line_end = line_start + line.count("\n")
            heading = line.lstrip("# ").strip() if line.startswith("#") else block_heading
            spans.append(
                ParsedSpan(
                    span_id=f"span_{short_hash(f'{source_id}:{start}:{end}', length=10)}",
                    source_id=source_id,
                    segment_kind="line",
                    heading=heading or None,
                    text=line,
                    span_start=start,
                    span_end=end,
                    line_start=line_start,
                    line_end=line_end,
                )
            )
            inner_cursor = end
        cursor = block_start + len(block)

    return spans


def datetime_or_none(timestamp: float | None) -> str | None:
    if timestamp is None:
        return None
    return now_utc().fromtimestamp(timestamp, tz=now_utc().tzinfo).isoformat()


def _parse_structured_json_source(
    *,
    payload: object,
    path: Path,
    project_relative_path: Path,
    source_id: str,
) -> ParsedSourceContent | None:
    lowered_path = project_relative_path.as_posix().lower()
    filename = path.name.lower()

    if _looks_like_ai_conversation_export(payload=payload, lowered_path=lowered_path, filename=filename):
        return _parse_ai_conversation_export(payload=payload, source_id=source_id)
    if _looks_like_reddit_export(payload=payload, lowered_path=lowered_path, filename=filename):
        return _parse_reddit_export(payload=payload, source_id=source_id)
    if _looks_like_google_activity_export(payload=payload, lowered_path=lowered_path, filename=filename):
        return _parse_google_activity_export(payload=payload, source_id=source_id)
    return None


def _looks_like_ai_conversation_export(*, payload: object, lowered_path: str, filename: str) -> bool:
    if (
        any(token in lowered_path or token in filename for token in ("chatgpt", "claude", "gemini", "assistant", "conversation"))
        and isinstance(payload, dict)
        and isinstance(payload.get("messages"), list)
    ):
        return True
    if isinstance(payload, dict) and isinstance(payload.get("messages"), list):
        return any(_coerce_message_text(message.get("content")) for message in payload["messages"] if isinstance(message, dict))
    return False


def _looks_like_reddit_export(*, payload: object, lowered_path: str, filename: str) -> bool:
    if (
        any(token in lowered_path or token in filename for token in ("reddit", "subreddit"))
        and isinstance(payload, dict)
        and any(isinstance(payload.get(key), list) for key in ("posts", "comments"))
    ):
        return True
    if not isinstance(payload, dict):
        return False
    posts = payload.get("posts")
    comments = payload.get("comments")
    return any(
        isinstance(item, dict) and item.get("subreddit") and (item.get("title") or item.get("body"))
        for collection in (posts, comments)
        if isinstance(collection, list)
        for item in collection
    )


def _looks_like_google_activity_export(*, payload: object, lowered_path: str, filename: str) -> bool:
    if (
        any(token in lowered_path or token in filename for token in ("google", "takeout", "activity", "search-history"))
        and isinstance(payload, list)
    ):
        return any(isinstance(item, dict) and (item.get("header") or item.get("titleUrl")) for item in payload)
    return isinstance(payload, list) and any(
        isinstance(item, dict) and item.get("time") and (item.get("header") or item.get("titleUrl"))
        for item in payload
    )


def _parse_ai_conversation_export(*, payload: object, source_id: str) -> ParsedSourceContent:
    if not isinstance(payload, dict):
        raise ValueError(
            f"AI conversation payload for {source_id} must be an object, got {type(payload).__name__}."
        )
    messages = payload.get("messages")
    if not isinstance(messages, list):
        raise ValueError(f"AI conversation payload for {source_id} must contain a messages list.")

    entries: list[dict[str, str]] = []
    timestamps: list[datetime] = []
    for message in messages:
        if not isinstance(message, dict):
            continue
        content = _coerce_message_text(message.get("content"))
        if not content:
            continue
        role = str(message.get("role") or message.get("author") or "unknown").strip().lower()
        entries.append(
            {
                "heading": role,
                "text": f"{role.title()}: {content}",
            }
        )
        parsed_timestamp = _coerce_datetime(message.get("created_at") or message.get("timestamp"))
        if parsed_timestamp:
            timestamps.append(parsed_timestamp)

    text, spans = _build_structured_spans(entries=entries, source_id=source_id)
    return ParsedSourceContent(
        text=text,
        spans=spans,
        source_type=SourceType.CHAT,
        source_category=SourceCategory.AI_DIALOGUE,
        parser="ai-conversation-json",
        title=str(payload.get("title") or "AI Conversation").strip(),
        created_at=max(timestamps) if timestamps else None,
        metadata={
            "structured_export_kind": "ai_conversation",
            "message_count": len(entries),
        },
        attachments=[],
    )


def _parse_agent_log_text(
    *,
    text: str,
    source_id: str,
    project_relative_path: Path,
    source_category: SourceCategory,
) -> tuple[str, list[ParsedSpan], dict[str, Any]] | None:
    lowered_path = project_relative_path.as_posix().lower()
    if "agent-logs" not in lowered_path and source_category != SourceCategory.AI_DIALOGUE:
        return None

    entries = _parse_jsonl_agent_log_entries(text)
    parser_kind = "agent-log-jsonl"
    if not entries:
        entries = _parse_markdown_agent_log_entries(text)
        parser_kind = "agent-log-markdown"
    if not entries:
        return None

    normalized_entries = [
        {"heading": entry["role"], "text": f"{entry['role'].title()}: {entry['content']}"}
        for entry in entries
        if entry["content"].strip()
    ]
    if not normalized_entries:
        return None
    parsed_text, spans = _build_structured_spans(entries=normalized_entries, source_id=source_id)
    role_counts: dict[str, int] = {}
    for entry in normalized_entries:
        role = entry["heading"]
        role_counts[role] = role_counts.get(role, 0) + 1
    return (
        parsed_text,
        spans,
        {
            "structured_export_kind": parser_kind,
            "message_count": len(normalized_entries),
            "role_counts": role_counts,
            "attribution_policy": "only user/human/developer spans may produce person-skill evidence",
        },
    )


def _parse_jsonl_agent_log_entries(text: str) -> list[dict[str, str]]:
    entries: list[dict[str, str]] = []
    parsed_lines = 0
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line or not line.startswith("{"):
            continue
        try:
            payload = json.loads(line)
        except json.JSONDecodeError:
            continue
        if not isinstance(payload, dict):
            continue
        role = _normalize_agent_log_role(payload.get("role") or payload.get("type") or payload.get("speaker"))
        content = _coerce_message_text(
            payload.get("content")
            or payload.get("message")
            or payload.get("text")
            or payload.get("markdown")
            or payload.get("output")
        )
        if role and content:
            entries.append({"role": role, "content": content})
            parsed_lines += 1
    # Avoid treating arbitrary JSONL data as chat unless it has multiple role-tagged records.
    return entries if parsed_lines >= 2 else []


def _parse_markdown_agent_log_entries(text: str) -> list[dict[str, str]]:
    marker = re.compile(
        r"^\s{0,3}(?:#{1,6}\s*)?"
        r"(?:\*\*)?"
        r"(user|human|developer|assistant|ai|agent|model|system|tool|function|thinking|reasoning|analysis)"
        r"(?:\*\*)?"
        r"\s*(?::|-|$)",
        re.IGNORECASE,
    )
    entries: list[dict[str, str]] = []
    current_role: str | None = None
    current_lines: list[str] = []

    def flush() -> None:
        nonlocal current_role, current_lines
        if current_role and any(line.strip() for line in current_lines):
            entries.append({"role": current_role, "content": "\n".join(current_lines).strip()})
        current_role = None
        current_lines = []

    for line in text.splitlines():
        match = marker.match(line)
        if match:
            flush()
            current_role = _normalize_agent_log_role(match.group(1))
            remainder = line[match.end() :].strip()
            if remainder:
                current_lines.append(remainder)
            continue
        if current_role:
            current_lines.append(line)
    flush()
    role_set = {entry["role"] for entry in entries}
    if role_set & USER_AGENT_LOG_ROLES and role_set & NON_USER_AGENT_LOG_ROLES:
        return entries
    return []


def _normalize_agent_log_role(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    normalized = value.strip().lower().replace("_", "-")
    if normalized in USER_AGENT_LOG_ROLES:
        return "user"
    if normalized in NON_USER_AGENT_LOG_ROLES:
        if normalized in {"thinking", "reasoning", "analysis"}:
            return "thinking"
        if normalized in {"tool", "function"}:
            return "tool"
        if normalized == "system":
            return "system"
        return "assistant"
    return None


def _parse_reddit_export(*, payload: object, source_id: str) -> ParsedSourceContent:
    if not isinstance(payload, dict):
        raise ValueError("Reddit export payload must be an object")

    entries: list[dict[str, str]] = []
    timestamps: list[datetime] = []
    post_count = 0
    comment_count = 0

    for post in payload.get("posts", []):
        if not isinstance(post, dict):
            continue
        subreddit = str(post.get("subreddit") or "unknown").strip()
        title = str(post.get("title") or "").strip()
        body = str(post.get("selftext") or post.get("body") or "").strip()
        if not title and not body:
            continue
        entries.append(
            {
                "heading": f"r/{subreddit}",
                "text": _join_parts(
                    [
                        f"Reddit post in r/{subreddit}",
                        title,
                        body,
                    ]
                ),
            }
        )
        post_count += 1
        parsed_timestamp = _coerce_datetime(post.get("created_at")) or _coerce_unix_datetime(post.get("created_utc"))
        if parsed_timestamp:
            timestamps.append(parsed_timestamp)

    for comment in payload.get("comments", []):
        if not isinstance(comment, dict):
            continue
        subreddit = str(comment.get("subreddit") or "unknown").strip()
        body = str(comment.get("body") or "").strip()
        if not body:
            continue
        entries.append(
            {
                "heading": f"r/{subreddit}",
                "text": _join_parts(
                    [
                        f"Reddit comment in r/{subreddit}",
                        body,
                    ]
                ),
            }
        )
        comment_count += 1
        parsed_timestamp = _coerce_datetime(comment.get("created_at")) or _coerce_unix_datetime(comment.get("created_utc"))
        if parsed_timestamp:
            timestamps.append(parsed_timestamp)

    text, spans = _build_structured_spans(entries=entries, source_id=source_id)
    return ParsedSourceContent(
        text=text,
        spans=spans,
        source_type=SourceType.JSON,
        source_category=SourceCategory.SOCIAL_OR_COMMUNITY_TRACE,
        parser="reddit-export-json",
        title="Reddit Export",
        created_at=max(timestamps) if timestamps else None,
        metadata={
            "structured_export_kind": "reddit_activity",
            "post_count": post_count,
            "comment_count": comment_count,
        },
        attachments=[],
    )


def _parse_google_activity_export(*, payload: object, source_id: str) -> ParsedSourceContent:
    if not isinstance(payload, list):
        raise ValueError("Google activity payload must be a list")

    entries: list[dict[str, str]] = []
    timestamps: list[datetime] = []
    for item in payload:
        if not isinstance(item, dict):
            continue
        header = str(item.get("header") or "activity").strip()
        title = str(item.get("title") or "").strip()
        url = str(item.get("titleUrl") or item.get("url") or "").strip()
        if not title and not url:
            continue
        entries.append(
            {
                "heading": header.lower(),
                "text": _join_parts([header, title, url]),
            }
        )
        parsed_timestamp = _coerce_datetime(item.get("time"))
        if parsed_timestamp:
            timestamps.append(parsed_timestamp)

    text, spans = _build_structured_spans(entries=entries, source_id=source_id)
    return ParsedSourceContent(
        text=text,
        spans=spans,
        source_type=SourceType.JSON,
        source_category=SourceCategory.PLATFORM_EXPORT_ACTIVITY,
        parser="google-activity-json",
        title="Google Activity Export",
        created_at=max(timestamps) if timestamps else None,
        metadata={
            "structured_export_kind": "google_activity",
            "activity_count": len(entries),
        },
        attachments=[],
    )


def _build_structured_spans(*, entries: list[dict[str, str]], source_id: str) -> tuple[str, list[ParsedSpan]]:
    text_parts: list[str] = []
    spans: list[ParsedSpan] = []
    cursor = 0
    line_cursor = 1
    for index, entry in enumerate(entries):
        entry_text = entry["text"]
        if not entry_text:
            continue
        if text_parts:
            text_parts.append("\n\n")
            cursor += 2
            line_cursor += 2
        start = cursor
        end = start + len(entry_text)
        line_start = line_cursor
        line_end = line_start + entry_text.count("\n")
        spans.append(
            ParsedSpan(
                span_id=f"span_{short_hash(f'{source_id}:structured:{index}:{start}:{end}', length=10)}",
                source_id=source_id,
                segment_kind="structured_entry",
                heading=entry.get("heading") or None,
                text=entry_text,
                span_start=start,
                span_end=end,
                line_start=line_start,
                line_end=line_end,
            )
        )
        text_parts.append(entry_text)
        cursor = end
        line_cursor = line_end
    return "".join(text_parts), spans


def _coerce_message_text(content: object) -> str:
    if isinstance(content, str):
        return content.strip()
    if isinstance(content, list):
        parts: list[str] = []
        for item in content:
            if isinstance(item, str) and item.strip():
                parts.append(item.strip())
            elif isinstance(item, dict):
                text_value = item.get("text") or item.get("content")
                if isinstance(text_value, str) and text_value.strip():
                    parts.append(text_value.strip())
        return " ".join(parts).strip()
    if isinstance(content, dict):
        for key in ("text", "content", "body", "value"):
            value = content.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip()
        if isinstance(content.get("parts"), list):
            return _coerce_message_text(content["parts"])
    return ""


def _coerce_datetime(value: object) -> datetime | None:
    if not isinstance(value, str) or not value.strip():
        return None
    raw_value = value.strip()
    if raw_value.endswith("Z"):
        raw_value = raw_value.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(raw_value)
    except ValueError:
        return None
    return parsed if parsed.tzinfo else parsed.replace(tzinfo=UTC)


def _coerce_unix_datetime(value: object) -> datetime | None:
    if isinstance(value, str) and value.isdigit():
        value = int(value)
    if not isinstance(value, int | float):
        return None
    return datetime.fromtimestamp(float(value), tz=UTC)


def _join_parts(parts: list[str]) -> str:
    return ". ".join(part.strip() for part in parts if part and part.strip())


def _classify_source_category(
    *,
    path: Path,
    project_relative_path: Path,
    source_type: SourceType,
    text: str,
) -> SourceCategory:
    del text
    lowered_path = project_relative_path.as_posix().lower()
    filename = path.name.lower()

    if "agent-logs" in lowered_path:
        return SourceCategory.AI_DIALOGUE
    if any(token in lowered_path or token in filename for token in ("chatgpt", "claude", "gemini", "assistant", "ai-chat", "conversation")):
        return SourceCategory.AI_DIALOGUE
    if any(token in lowered_path or token in filename for token in ("reddit", "twitter", "tweet", "x-", "mastodon", "forum", "discord", "slack", "social", "profile", "linkedin")):
        return SourceCategory.SOCIAL_OR_COMMUNITY_TRACE
    if source_type in {
        SourceType.MARKDOWN,
        SourceType.TEXT,
        SourceType.DOCX,
        SourceType.PDF,
        SourceType.SPREADSHEET,
    } and any(token in lowered_path for token in ("/drive/", "drive/")):
        return SourceCategory.AUTHORED_CONTENT
    if any(token in lowered_path or token in filename for token in ("takeout", "export", "search", "history", "activity", "browser", "watch-history", "bookmarks")):
        return SourceCategory.PLATFORM_EXPORT_ACTIVITY
    if any(token in lowered_path or token in filename for token in ("issue", "ticket", "review", "pull-request", "pr-", "standup", "retro", "meeting")):
        return SourceCategory.COLLABORATION_TRACE
    if source_type == SourceType.CODE:
        return SourceCategory.PRODUCED_ARTIFACT
    if source_type in {SourceType.JSON, SourceType.CSV}:
        return SourceCategory.METADATA_ONLY_ACTIVITY
    if source_type in {
        SourceType.MARKDOWN,
        SourceType.TEXT,
        SourceType.DOCX,
        SourceType.PDF,
        SourceType.SPREADSHEET,
    }:
        return SourceCategory.AUTHORED_CONTENT
    return SourceCategory.AUTHORED_CONTENT


def _classification_reason(
    *,
    path: Path,
    project_relative_path: Path,
    source_category: SourceCategory,
) -> str:
    return (
        f"classified_as={source_category.value} "
        f"from path={project_relative_path.as_posix()} filename={path.name}"
    )


_ATTACHMENT_SUFFIX_TO_KIND = {
    ".jpg": AttachmentKind.IMAGE,
    ".jpeg": AttachmentKind.IMAGE,
    ".png": AttachmentKind.IMAGE,
    ".gif": AttachmentKind.IMAGE,
    ".webp": AttachmentKind.IMAGE,
    ".bmp": AttachmentKind.IMAGE,
    ".svg": AttachmentKind.IMAGE,
    ".avif": AttachmentKind.IMAGE,
    ".heic": AttachmentKind.IMAGE,
    ".heif": AttachmentKind.IMAGE,
    ".mp4": AttachmentKind.VIDEO,
    ".mov": AttachmentKind.VIDEO,
    ".mkv": AttachmentKind.VIDEO,
    ".webm": AttachmentKind.VIDEO,
    ".avi": AttachmentKind.VIDEO,
    ".mp3": AttachmentKind.AUDIO,
    ".wav": AttachmentKind.AUDIO,
    ".m4a": AttachmentKind.AUDIO,
    ".ogg": AttachmentKind.AUDIO,
    ".flac": AttachmentKind.AUDIO,
    ".aac": AttachmentKind.AUDIO,
    ".pdf": AttachmentKind.DOCUMENT,
}

_ATTACHMENT_LIKE_KEYS = (
    "image",
    "img",
    "photo",
    "video",
    "audio",
    "media",
    "thumbnail",
    "thumb",
    "preview",
    "poster",
    "attachment",
    "asset",
    "icon",
    "cover",
    "banner",
    "src",
)

_REMOTE_MEDIA_URL_PATTERN = re.compile(r"https?://[^\s<>'\")\]]+")
_REMOTE_MEDIA_HOSTS = {
    "youtube.com",
    "www.youtube.com",
    "m.youtube.com",
    "music.youtube.com",
    "youtu.be",
    "www.youtu.be",
    "youtube-nocookie.com",
    "www.youtube-nocookie.com",
}
_REMOTE_MEDIA_ATTACHMENT_KINDS = {
    AttachmentKind.VIDEO,
    AttachmentKind.AUDIO,
}


def _collect_json_attachment_references(value: object, *, key_hint: str | None = None) -> list[dict[str, Any]]:
    references: list[dict[str, Any]] = []
    if isinstance(value, dict):
        for key, nested_value in value.items():
            references.extend(_collect_json_attachment_references(nested_value, key_hint=str(key)))
        return references
    if isinstance(value, list):
        for item in value:
            references.extend(_collect_json_attachment_references(item, key_hint=key_hint))
        return references

    if key_hint and not any(token in key_hint.lower() for token in _ATTACHMENT_LIKE_KEYS):
        return []

    reference = _attachment_reference_from_value(value, contextual_hint=key_hint)
    return [reference] if reference else []


def _attachment_reference_from_value(
    value: object,
    *,
    label: str | None = None,
    contextual_hint: str | None = None,
) -> dict[str, Any] | None:
    if not isinstance(value, str):
        return None
    cleaned = value.strip()
    if not cleaned:
        return None
    parsed = _safe_urlparse(cleaned)
    if parsed is None:
        return None
    suffix = Path(parsed.path or cleaned).suffix.lower()
    kind = _ATTACHMENT_SUFFIX_TO_KIND.get(suffix)
    if kind is None:
        return None
    payload: dict[str, Any] = {
        "reference": cleaned,
        "kind": kind.value,
    }
    if label:
        payload["label"] = label
    if contextual_hint:
        payload["contextual_hint"] = contextual_hint
    return payload


def _collect_remote_media_url_references_from_text(text: str) -> list[dict[str, Any]]:
    references: list[dict[str, Any]] = []
    for match in _REMOTE_MEDIA_URL_PATTERN.finditer(text):
        reference = _remote_media_reference_from_url(
            match.group(0),
            contextual_hint="inline_media_url",
        )
        if reference:
            references.append(reference)
    return references


def _remote_media_reference_from_url(
    value: str,
    *,
    contextual_hint: str | None = None,
) -> dict[str, Any] | None:
    cleaned = value.rstrip(".,;:!?)]}\"'")
    parsed = _safe_urlparse(cleaned)
    if parsed is None:
        return None
    if parsed.scheme not in {"http", "https"}:
        return None

    host = parsed.netloc.lower().split("@")[-1].split(":", 1)[0]
    if host in _REMOTE_MEDIA_HOSTS:
        kind = AttachmentKind.VIDEO
    else:
        suffix = Path(parsed.path).suffix.lower()
        kind = _ATTACHMENT_SUFFIX_TO_KIND.get(suffix)
        if kind not in _REMOTE_MEDIA_ATTACHMENT_KINDS:
            return None

    payload: dict[str, Any] = {
        "reference": cleaned,
        "kind": kind.value,
    }
    if contextual_hint:
        payload["contextual_hint"] = contextual_hint
    return payload


def _build_linked_attachments(
    *,
    source_path: Path,
    attachment_references: list[dict[str, Any]],
    config: TracciaConfig | None,
) -> list[SourceAttachment]:
    if config and not config.multimodal.enable_linked_attachments:
        return []

    max_attachments = config.multimodal.max_attachments_per_source if config else 4
    max_attachment_text_characters = (
        config.multimodal.max_attachment_text_characters if config else 1200
    )
    max_attachment_transcript_characters = (
        config.multimodal.max_attachment_transcript_characters if config else 8000
    )
    enable_local_image_ocr = config.multimodal.enable_local_image_ocr if config else True
    enable_local_media_transcription = (
        config.multimodal.enable_local_media_transcription if config else True
    )
    transcription_provider = (
        config.multimodal.audio_transcription_provider if config else "auto"
    )
    transcription_model = (
        config.multimodal.audio_transcription_model if config else "turbo"
    )
    transcription_device = (
        config.multimodal.audio_transcription_device if config else "cpu"
    )
    ocr_timeout_seconds = config.multimodal.ocr_timeout_seconds if config else 20
    transcription_timeout_seconds = (
        config.multimodal.transcription_timeout_seconds if config else 1800
    )
    enable_remote_media_enrichment = (
        config.multimodal.enable_remote_media_enrichment if config else True
    )
    remote_media_enrichment_command = (
        config.multimodal.remote_media_enrichment_command if config else "summarize"
    )
    remote_media_enrichment_video_mode = (
        config.multimodal.remote_media_enrichment_video_mode if config else "understand"
    )
    enable_remote_media_slides = (
        config.multimodal.enable_remote_media_slides if config else True
    )
    enable_remote_media_slides_ocr = (
        config.multimodal.enable_remote_media_slides_ocr if config else True
    )
    remote_media_enrichment_timeout_seconds = (
        config.multimodal.remote_media_enrichment_timeout_seconds if config else 180
    )

    attachments: list[SourceAttachment] = []
    seen: set[tuple[str, str]] = set()

    for reference_payload in attachment_references:
        if len(attachments) >= max_attachments:
            break
        reference = str(reference_payload.get("reference") or "").strip()
        kind_value = str(reference_payload.get("kind") or "").strip()
        if not reference or not kind_value:
            continue
        dedupe_key = (kind_value, reference)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)

        kind = AttachmentKind(kind_value)
        resolved_path = _resolve_attachment_path(source_path=source_path, reference=reference)
        uri = None
        mime_type = None
        extracted_text = None
        attachment_metadata: dict[str, Any] = {}
        if resolved_path:
            uri = resolved_path.as_uri()
            mime_type = mimetypes.guess_type(resolved_path.name)[0]
            if kind == AttachmentKind.IMAGE and enable_local_image_ocr:
                extracted_text = _extract_image_text(
                    resolved_path,
                    max_characters=max_attachment_text_characters,
                    timeout_seconds=ocr_timeout_seconds,
                )
            elif kind in {AttachmentKind.AUDIO, AttachmentKind.VIDEO} and enable_local_media_transcription:
                extracted_text, attachment_metadata = _extract_media_text(
                    resolved_path,
                    kind=kind,
                    provider=transcription_provider,
                    model=transcription_model,
                    device=transcription_device,
                    max_characters=max_attachment_transcript_characters,
                    timeout_seconds=transcription_timeout_seconds,
                )
        elif reference.startswith(("http://", "https://", "file://")):
            uri = reference
            mime_type = mimetypes.guess_type(reference)[0]
            if enable_remote_media_enrichment and kind in _REMOTE_MEDIA_ATTACHMENT_KINDS:
                extracted_text, attachment_metadata = _extract_remote_media_text(
                    reference,
                    command=remote_media_enrichment_command,
                    video_mode=remote_media_enrichment_video_mode,
                    enable_slides=enable_remote_media_slides,
                    enable_slides_ocr=enable_remote_media_slides_ocr,
                    max_characters=max_attachment_transcript_characters,
                    timeout_seconds=remote_media_enrichment_timeout_seconds,
                )

        attachments.append(
            SourceAttachment(
                attachment_id=f"att_{short_hash(f'{source_path}:{reference}:{kind.value}', length=12)}",
                kind=kind,
                reference=reference,
                resolved_path=resolved_path.as_posix() if resolved_path else None,
                uri=uri,
                mime_type=mime_type,
                label=str(reference_payload.get("label") or "").strip() or None,
                extracted_text=extracted_text,
                contextual_hint=str(reference_payload.get("contextual_hint") or "").strip() or None,
                metadata=attachment_metadata,
            )
        )

    return attachments


def _resolve_attachment_path(*, source_path: Path, reference: str) -> Path | None:
    if reference.startswith(("http://", "https://", "data:")):
        return None
    if reference.startswith("file://"):
        candidate = Path(reference.removeprefix("file://"))
        try:
            return candidate.resolve() if candidate.exists() else None
        except OSError:
            return None

    parsed = _safe_urlparse(reference)
    raw_path = (parsed.path if parsed else None) or reference
    if not raw_path:
        return None

    candidate = Path(raw_path)
    try:
        if not candidate.is_absolute():
            relative_candidate = candidate
            candidate = (source_path.parent / relative_candidate).resolve()
            if not candidate.exists() and source_path.is_symlink():
                candidate = (source_path.resolve().parent / relative_candidate).resolve()
        else:
            candidate = candidate.resolve()
        return candidate if candidate.exists() else None
    except OSError:
        # Export rows can contain arbitrary user/message text in fields that look
        # attachment-like. Some of that text is longer than filesystem component
        # limits; it is not a local attachment and must not fail ingestion.
        return None


def _safe_urlparse(value: str) -> ParseResult | None:
    try:
        return urlparse(value)
    except ValueError:
        return None


def _extract_image_text(path: Path, *, max_characters: int, timeout_seconds: int) -> str | None:
    if path.suffix.lower() == ".svg":
        return None
    try:
        completed = subprocess.run(
            ["tesseract", str(path), "stdout", "--psm", "6"],
            capture_output=True,
            text=True,
            timeout=timeout_seconds,
            check=False,
        )
    except (FileNotFoundError, subprocess.SubprocessError, TimeoutError):
        return None
    if completed.returncode != 0:
        return None
    normalized = " ".join(completed.stdout.split())
    if not normalized:
        return None
    return normalized[:max_characters]


def _extract_media_text(
    path: Path,
    *,
    kind: AttachmentKind,
    provider: str,
    model: str,
    device: str,
    max_characters: int,
    timeout_seconds: int,
) -> tuple[str | None, dict[str, Any]]:
    resolved_provider = _resolve_media_transcription_provider(provider)
    if resolved_provider is None:
        return None, {}
    if shutil.which("ffmpeg") is None or shutil.which("whisper") is None:
        return None, {}

    with tempfile.TemporaryDirectory(prefix="traccia-media-") as temp_dir:
        temp_root = Path(temp_dir)
        normalized_audio_path = temp_root / "attachment.wav"
        if not _normalize_media_audio(
            source_path=path,
            output_path=normalized_audio_path,
            kind=kind,
            timeout_seconds=timeout_seconds,
        ):
            return None, {}

        transcript = _transcribe_audio_with_whisper(
            normalized_audio_path,
            model=model,
            device=device,
            timeout_seconds=timeout_seconds,
        )
        if not transcript:
            return None, {}
        normalized = " ".join(transcript.split())
        if not normalized:
            return None, {}
        return (
            normalized[:max_characters],
            {
                "transcription_provider": resolved_provider,
                "transcription_model": model,
                "transcription_device": device,
            },
        )


def _extract_remote_media_text(
    reference: str,
    *,
    command: str,
    video_mode: str,
    enable_slides: bool,
    enable_slides_ocr: bool,
    max_characters: int,
    timeout_seconds: int,
) -> tuple[str | None, dict[str, Any]]:
    executable = command.strip()
    if not executable:
        return None, {}

    if "/" in executable:
        if not Path(executable).exists():
            return None, {}
        executable_label = Path(executable).name
    else:
        if shutil.which(executable) is None:
            return None, {}
        executable_label = executable

    normalized_video_mode = video_mode.strip().lower()
    if normalized_video_mode not in {"auto", "transcript", "understand"}:
        raise ValueError(
            "Unsupported multimodal.remote_media_enrichment_video_mode: "
            f"{video_mode}"
        )

    process_command = [
        executable,
        reference,
        "--extract",
        "--plain",
        "--format",
        "text",
        "--video-mode",
        normalized_video_mode,
        "--youtube",
        "auto",
        "--timeout",
        f"{timeout_seconds}s",
        "--max-extract-characters",
        str(max_characters),
    ]
    if enable_slides:
        process_command.append("--slides")
    if enable_slides_ocr:
        process_command.append("--slides-ocr")
    try:
        completed = subprocess.run(
            process_command,
            capture_output=True,
            text=True,
            timeout=timeout_seconds + 15,
            check=False,
        )
    except (FileNotFoundError, subprocess.SubprocessError, TimeoutError):
        return None, {}
    if completed.returncode != 0:
        return None, {}

    normalized = " ".join(completed.stdout.split())
    if not normalized:
        return None, {}
    return (
        normalized[:max_characters],
        {
            "url_enrichment_provider": "summarize_cli",
            "url_enrichment_command": executable_label,
            "url_enrichment_mode": "extract",
            "url_enrichment_video_mode": normalized_video_mode,
            "url_enrichment_slides": enable_slides,
            "url_enrichment_slides_ocr": enable_slides_ocr,
        },
    )


def _resolve_media_transcription_provider(provider: str) -> str | None:
    normalized = provider.strip().lower()
    if normalized in {"none", "disabled"}:
        return None
    if normalized == "auto":
        return "whisper_cli" if shutil.which("whisper") else None
    if normalized in {"whisper", "whisper_cli"}:
        return "whisper_cli"
    raise ValueError(f"Unsupported multimodal.audio_transcription_provider: {provider}")


def _normalize_media_audio(
    *,
    source_path: Path,
    output_path: Path,
    kind: AttachmentKind,
    timeout_seconds: int,
) -> bool:
    command = [
        "ffmpeg",
        "-nostdin",
        "-y",
        "-i",
        str(source_path),
    ]
    if kind == AttachmentKind.VIDEO:
        command.append("-vn")
    command.extend(
        [
            "-ac",
            "1",
            "-ar",
            "16000",
            str(output_path),
        ]
    )
    try:
        completed = subprocess.run(
            command,
            capture_output=True,
            text=True,
            timeout=timeout_seconds,
            check=False,
        )
    except (FileNotFoundError, subprocess.SubprocessError, TimeoutError):
        return False
    return completed.returncode == 0 and output_path.exists()


def _transcribe_audio_with_whisper(
    audio_path: Path,
    *,
    model: str,
    device: str,
    timeout_seconds: int,
) -> str | None:
    output_dir = audio_path.parent
    command = [
        "whisper",
        str(audio_path),
        "--model",
        model,
        "--device",
        device,
        "--output_dir",
        str(output_dir),
        "--output_format",
        "json",
        "--task",
        "transcribe",
        "--verbose",
        "False",
    ]
    if device.strip().lower() == "cpu":
        command.extend(["--fp16", "False"])
    try:
        completed = subprocess.run(
            command,
            capture_output=True,
            text=True,
            timeout=timeout_seconds,
            check=False,
        )
    except (FileNotFoundError, subprocess.SubprocessError, TimeoutError):
        return None
    if completed.returncode != 0:
        return None
    transcript_path = output_dir / f"{audio_path.stem}.json"
    if not transcript_path.exists():
        return None
    try:
        payload = json.loads(transcript_path.read_text())
    except (OSError, json.JSONDecodeError):
        return None
    text = payload.get("text")
    return text.strip() if isinstance(text, str) and text.strip() else None
